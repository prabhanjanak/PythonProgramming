from openpyxl import load_workbook
def search_sheet(sheet_name, search_key, search_column, result_column):
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
        if row[search_column] == search_key:
            print(f"Corresponding {sheet_name.lower()} for code {search_key} is {row[result_column]}")
            return
    print(f"State code {search_key} not found in the '{sheet_name}' sheet.")
# Load the existing workbook
wb = load_workbook("demo.xlsx")
search_sheet("Capital", input("Enter state code for finding capital: "), 2, 1)
search_sheet("Language", input("Enter state code for finding language: "), 2, 1)
wb.close()
